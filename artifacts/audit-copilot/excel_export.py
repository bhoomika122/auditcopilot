import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

NAVY = "0B2545"
GOLD = "C9A961"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
RED_BG = "FCE4E4"
AMBER_BG = "FEF3CD"
GREEN_BG = "D4EDDA"


def _navy_fill():
    return PatternFill("solid", fgColor=NAVY)


def _gold_fill():
    return PatternFill("solid", fgColor=GOLD)


def _header_font():
    return Font(bold=True, color=WHITE, name="Calibri", size=11)


def _title_font():
    return Font(bold=True, color=NAVY, name="Calibri", size=14)


def _body_font(bold=False):
    return Font(bold=bold, name="Calibri", size=10)


def _thin_border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def _risk_fill(risk_level: str):
    mapping = {
        "High": PatternFill("solid", fgColor="FADBD8"),
        "Medium": PatternFill("solid", fgColor="FDEBD0"),
        "Low": PatternFill("solid", fgColor="D5F5E3"),
    }
    return mapping.get(risk_level, PatternFill("solid", fgColor=LIGHT_GRAY))


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def _write_summary_sheet(ws, df_all: pd.DataFrame, df_flagged: pd.DataFrame):
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "AuditCopilot — Workpaper Summary"
    cell.font = _title_font()
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws["A2"].value = f"Generated: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = _body_font()
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.append([])

    kpi_headers = ["Metric", "Value"]
    ws.append(kpi_headers)
    for i, h in enumerate(kpi_headers, 1):
        cell = ws.cell(row=ws.max_row, column=i)
        cell.fill = _navy_fill()
        cell.font = _header_font()
        cell.alignment = Alignment(horizontal="center")
        cell.border = _thin_border()

    kpis = [
        ("Total Transactions Reviewed", len(df_all)),
        ("Total Flagged Transactions", len(df_flagged)),
        ("Unique Vendors Flagged", df_flagged["vendor_name"].nunique() if not df_flagged.empty else 0),
        ("Total Dollar Value at Risk", f"${df_flagged['amount'].sum():,.2f}" if not df_flagged.empty else "$0.00"),
        ("High Risk Flags", len(df_flagged[df_flagged["risk_level"] == "High"]) if not df_flagged.empty else 0),
        ("Medium Risk Flags", len(df_flagged[df_flagged["risk_level"] == "Medium"]) if not df_flagged.empty else 0),
        ("Low Risk Flags", len(df_flagged[df_flagged["risk_level"] == "Low"]) if not df_flagged.empty else 0),
    ]

    for metric, value in kpis:
        ws.append([metric, value])
        for i in range(1, 3):
            cell = ws.cell(row=ws.max_row, column=i)
            cell.font = _body_font()
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal="left" if i == 1 else "right")

    ws.append([])

    flag_cat_header = ["Flag Category", "Count", "$ Value at Risk", "Risk Level"]
    ws.append(flag_cat_header)
    for i, h in enumerate(flag_cat_header, 1):
        cell = ws.cell(row=ws.max_row, column=i)
        cell.fill = _gold_fill()
        cell.font = Font(bold=True, color=NAVY, name="Calibri", size=11)
        cell.border = _thin_border()

    if not df_flagged.empty:
        for cat, grp in df_flagged.groupby("flag_category"):
            risk = grp["risk_level"].mode()[0] if not grp.empty else "Medium"
            ws.append([cat, len(grp), f"${grp['amount'].sum():,.2f}", risk])
            for i in range(1, 5):
                cell = ws.cell(row=ws.max_row, column=i)
                cell.font = _body_font()
                cell.border = _thin_border()
                if i == 4:
                    cell.fill = _risk_fill(risk)

    _auto_width(ws)
    ws.freeze_panes = "A4"


def _write_flagged_sheet(ws, df_flagged: pd.DataFrame):
    ws.title = "Flagged Transactions"
    ws.sheet_view.showGridLines = False

    display_cols = [
        "transaction_id", "vendor_name", "invoice_number",
        "invoice_date", "posting_date", "amount", "currency",
        "gl_account", "flag_category", "risk_level", "rule_explanation"
    ]
    available = [c for c in display_cols if c in df_flagged.columns]
    df_out = df_flagged[available].copy() if not df_flagged.empty else pd.DataFrame(columns=available)

    header_row = [c.replace("_", " ").title() for c in available]
    ws.append(header_row)
    for i, h in enumerate(header_row, 1):
        cell = ws.cell(row=1, column=i)
        cell.fill = _navy_fill()
        cell.font = _header_font()
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = _thin_border()

    for row_data in df_out.itertuples(index=False):
        ws.append(list(row_data))
        row_idx = ws.max_row
        risk = str(row_data.risk_level) if hasattr(row_data, "risk_level") else "Medium"
        for i in range(1, len(available) + 1):
            cell = ws.cell(row=row_idx, column=i)
            cell.font = _body_font()
            cell.border = _thin_border()
            cell.fill = _risk_fill(risk)
            col_name = available[i - 1]
            if col_name == "amount":
                cell.number_format = '#,##0.00'

    _auto_width(ws)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30


def _write_memos_sheet(ws, df_flagged: pd.DataFrame, memos_dict: dict):
    ws.title = "AI Memos"
    ws.sheet_view.showGridLines = False

    headers = ["Transaction ID", "Vendor", "Flag Category", "Amount", "AI Audit Memo"]
    ws.append(headers)
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i)
        cell.fill = _navy_fill()
        cell.font = _header_font()
        cell.alignment = Alignment(horizontal="center")
        cell.border = _thin_border()

    if not df_flagged.empty:
        for _, row in df_flagged.iterrows():
            tid = row.get("transaction_id", "")
            memo = memos_dict.get(tid, "No memo generated.")
            ws.append([
                tid,
                row.get("vendor_name", ""),
                row.get("flag_category", ""),
                f"${row.get('amount', 0):,.2f}",
                memo,
            ])
            row_idx = ws.max_row
            for i in range(1, 6):
                cell = ws.cell(row=row_idx, column=i)
                cell.font = _body_font()
                cell.border = _thin_border()
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    for col_letter, width in [("A", 18), ("B", 30), ("C", 25), ("D", 15), ("E", 80)]:
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30


def build_workpaper(
    df_all: pd.DataFrame,
    df_flagged: pd.DataFrame,
    memos_dict: dict,
) -> io.BytesIO:
    wb = Workbook()
    ws_summary = wb.active
    ws_flagged = wb.create_sheet()
    ws_memos = wb.create_sheet()

    _write_summary_sheet(ws_summary, df_all, df_flagged)
    _write_flagged_sheet(ws_flagged, df_flagged)
    _write_memos_sheet(ws_memos, df_flagged, memos_dict)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
